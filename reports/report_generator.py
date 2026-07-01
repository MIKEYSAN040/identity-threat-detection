import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime

RISK_COLORS = {
    "CRITICAL": "C00000",
    "HIGH":     "FF4500",
    "MEDIUM":   "FFA500",
    "LOW":      "4CAF50",
}

HEADER_COLOR = "1F3864"

def thin_border():
    side = Side(style="thin", color="CCCCCC")
    return Border(left=side, right=side, top=side, bottom=side)

def generate_report(results, summary, output_path="reports/identity_threat_report.xlsx"):
    wb = openpyxl.Workbook()

    # ── Sheet 1: Executive Summary ──────────────────────────────────
    ws1 = wb.active
    ws1.title = "Executive Summary"
    ws1.column_dimensions["A"].width = 30
    ws1.column_dimensions["B"].width = 20

    ws1.merge_cells("A1:B1")
    ws1["A1"] = "Identity Threat Detection Report"
    ws1["A1"].font = Font(bold=True, size=16, color="FFFFFF")
    ws1["A1"].fill = PatternFill("solid", fgColor=HEADER_COLOR)
    ws1["A1"].alignment = Alignment(horizontal="center")

    ws1["A2"] = "Generated"
    ws1["B2"] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    ws1["A3"] = "Total Users Analyzed"
    ws1["B3"] = len(results)

    ws1["A4"] = "Total Alerts"
    ws1["B4"] = summary["total_alerts"]

    row = 6
    for level, color in RISK_COLORS.items():
        ws1[f"A{row}"] = f"{level} Risk Users"
        ws1[f"B{row}"] = summary.get(level, 0)
        ws1[f"A{row}"].fill = PatternFill("solid", fgColor=color)
        ws1[f"A{row}"].font = Font(bold=True, color="FFFFFF")
        row += 1

    # ── Sheet 2: Findings (detailed) ─────────────────────────────────
    ws2 = wb.create_sheet("Findings")
    headers = ["Username", "Risk Level", "Risk Score", "Detection Type",
               "Detail", "Event Time", "Source IP", "MITRE ID", "MITRE Technique", "Tactic"]
    col_widths = [15, 12, 12, 22, 55, 22, 18, 12, 28, 22]

    for i, (h, w) in enumerate(zip(headers, col_widths), 1):
        cell = ws2.cell(row=1, column=i, value=h)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor=HEADER_COLOR)
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
        cell.border = thin_border()
        ws2.column_dimensions[get_column_letter(i)].width = w

    ws2.row_dimensions[1].height = 30
    row = 2
    for user in results:
        for finding in user["findings"]:
            values = [
                user["username"],
                user["risk_level"],
                user["risk_score"],
                finding["type"].replace("_", " ").title(),
                finding["detail"],
                finding["eventTime"],
                finding["ip"],
                finding["mitre"]["id"],
                finding["mitre"]["name"],
                finding["mitre"]["tactic"],
            ]
            color = RISK_COLORS.get(user["risk_level"], "FFFFFF")
            for col, val in enumerate(values, 1):
                cell = ws2.cell(row=row, column=col, value=val)
                cell.border = thin_border()
                cell.alignment = Alignment(wrap_text=True, vertical="center")
                if col in (1, 2):
                    cell.fill = PatternFill("solid", fgColor=color)
                    cell.font = Font(bold=True, color="FFFFFF" if col == 2 else "000000")
            row += 1

    ws2.auto_filter.ref = f"A1:{get_column_letter(len(headers))}1"

    # ── Sheet 3: User Summary ─────────────────────────────────────────
    ws3 = wb.create_sheet("User Summary")
    h2 = ["Username", "Risk Level", "Risk Score", "Total Events", "Alert Count"]
    for i, h in enumerate(h2, 1):
        cell = ws3.cell(row=1, column=i, value=h)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor=HEADER_COLOR)
        cell.alignment = Alignment(horizontal="center")
        ws3.column_dimensions[get_column_letter(i)].width = 20

    for r, user in enumerate(results, 2):
        color = RISK_COLORS.get(user["risk_level"], "FFFFFF")
        vals = [user["username"], user["risk_level"], user["risk_score"],
                user["total_events"], user["finding_count"]]
        for c, v in enumerate(vals, 1):
            cell = ws3.cell(row=r, column=c, value=v)
            cell.border = thin_border()
            if c == 2:
                cell.fill = PatternFill("solid", fgColor=color)
                cell.font = Font(bold=True, color="FFFFFF")

    wb.save(output_path)
    print(f"Report saved to {output_path}")
    return output_path